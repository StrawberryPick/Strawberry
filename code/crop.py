import openpyxl
import random
import sqlite3

wb = openpyxl.load_workbook("StrawberryWeather.xlsx")
sheet = wb.active

cropdata = {    

}

rows = input("Enter how many rows you have in field  \n")
rows = int(rows)
day = input("Enter which day of the year you are interested  \n ")
fixday = int(day) + 1
#row_count = sheet.max_row
i = 1
thermal = sheet.cell(row=fixday, column=3).value
sunlight = wb.active.cell(row=fixday, column=4).value
wind = wb.active.cell(row=fixday, column=5).value
ultrav = wb.active.cell(row=fixday, column=6).value
thermal = float(thermal)
sunlight = float(sunlight)
wind = float(wind)
ultrav = float(ultrav)
while i <= rows:
    variance = random.uniform(1.1, 1.7)
    cropdata[i] = (10  * thermal * variance) + (sunlight*1.5) + (wind*1.1) + (ultrav*1.3)
    if cropdata[i] < 0:
        cropdata[i] = 0
    i = i + 1
#date = wb.active.cell(row=2, column=2).value
#print(date)
print(cropdata)
dbconnection = sqlite3.connect("StrawberryERP.db")
DBcursor = dbconnection.cursor()
for x in cropdata:
    InsertionTuple = (x, cropdata[x], day)
    DBcursor.execute("INSERT INTO CropPredictions (Row, Prednumber, PredDate) VALUES (?, ?, ?)", InsertionTuple)
    dbconnection.commit()
dbconnection.close()





